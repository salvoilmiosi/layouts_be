#!/usr/bin/env python

from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side
import sys
import json
import time
import os

class TableValue:
    def __init__(self, title, value, index=0, type='str', number_format='', column_width=None):
        self.title = title
        self.value = value
        self.index = index
        self.type = type
        self.number_format = number_format
        self.column_width = column_width

table_values = [
    TableValue('File',                  'filename'),
    TableValue('Ragione Sociale',       'ragione_sociale'),
    TableValue('Indirizzo Fornitura',   'indirizzo_fornitura'),
    TableValue('POD',                   'codice_pod',       column_width=16),
    TableValue('Mese',                  'mese_fattura',             type='month'),
    TableValue('Fornitore',             'fornitore'),
    TableValue('N. Fatt.',              'numero_fattura',   column_width=11),
    TableValue('Data Emissione',        'data_fattura',              type='date'),
    TableValue('Data scadenza',         'data_scadenza',             type='date'),
    TableValue('Costo Materia Energia', 'spesa_materia_energia',     type='euro', column_width=11),
    TableValue('Trasporto',             'trasporto_gestione',        type='euro', column_width=11),
    TableValue('Oneri',                 'oneri',                     type='euro', column_width=11),
    TableValue('Accise',                'accise',                    type='euro', column_width=11),
    TableValue('Iva',                   'iva',                       type='percentage', column_width=6),
    TableValue('Imponibile',            'imponibile',                type='euro', column_width=11),
    TableValue('F1 Rilev.',             'energia_attiva_rilevata', index=0, type='int', column_width=8),
    TableValue('F2 Rilev.',             'energia_attiva_rilevata', index=1, type='int', column_width=8),
    TableValue('F3 Rilev.',             'energia_attiva_rilevata', index=2, type='int', column_width=8),
    TableValue('F1',                    'energia_attiva',   index=0, type='int', column_width=8),
    TableValue('F2',                    'energia_attiva',   index=1, type='int', column_width=8),
    TableValue('F3',                    'energia_attiva',   index=2, type='int', column_width=8),
    TableValue('P1',                    'potenza',          index=0, type='int', column_width=8),
    TableValue('P2',                    'potenza',          index=1, type='int', column_width=8),
    TableValue('P3',                    'potenza',          index=2, type='int', column_width=8),
    TableValue('R1',                    'energia_reattiva', index=0, type='int', column_width=8),
    TableValue('R2',                    'energia_reattiva', index=1, type='int', column_width=8),
    TableValue('R3',                    'energia_reattiva', index=2, type='int', column_width=8),
    TableValue('Oneri Ammin.',          'oneri_amministrativi',      type='euro'),
    TableValue('PCV',                   'pcv',                       type='euro'),
    TableValue('CTS',                   'cts',                       type='euro'),
    TableValue('<75%',                  'penale_reattiva_inf75',     type='euro', column_width=11),
    TableValue('>75%',                  'penale_reattiva_sup75',     type='euro', column_width=11),
    TableValue('PEF1',                  'prezzo_energia',   index=0, type='number', number_format='0.00000000', column_width=11),
    TableValue('PEF2',                  'prezzo_energia',   index=1, type='number', number_format='0.00000000', column_width=11),
    TableValue('PEF3',                  'prezzo_energia',   index=2, type='number', number_format='0.00000000', column_width=11),
    TableValue('Sbilanciamento',        'sbilanciamento',            type='number', number_format='0.00000000', column_width=11),
    TableValue('Dispacciamento',        'dispacciamento',            type='number', number_format='0.00000000', column_width=11),
    TableValue('Disp. Var',             'disp_var',                  type='number', number_format='0.00000000', column_width=11)
]

def export_file(input_file):
    out = []
    out_err = []

    def add_rows(json_data):
        filename = json_data['filename']

        if 'error' in json_data:
            out_err.append([filename, json_data['error']])
            return
        
        for v in json_data['values']:
            row = []

            for obj in table_values:
                if obj.value == 'filename':
                    row.append({'value':filename, 'number_format': ''})
                else:
                    try:
                        value = v[obj.value][obj.index]
                        if obj.type == 'str':
                            row.append({'value': value, 'number_format': '@'})
                        elif obj.type == 'date':
                            row.append({'value': date.fromisoformat(value), 'number_format': 'DD/MM/YY'})
                        elif obj.type == 'month':
                            row.append({'value': date.fromisoformat(value), 'number_format': 'MM/YYYY'})
                        elif obj.type == 'euro':
                            row.append({'value': float(value), 'number_format': '#,##0.00 \u20ac'})
                        elif obj.type == 'int':
                            row.append({'value': float(value), 'number_format': '0'})
                        elif obj.type == 'number':
                            row.append({'value': float(value), 'number_format': obj.number_format})
                        elif obj.type == 'percentage':
                            row.append({'value': float(value[:-1]) / 100, 'number_format': '0%'})
                        else:
                            row.append({'value': value, 'number_format': ''})
                    except (KeyError, IndexError, ValueError):
                        row.append({'value': '', 'number_format': ''})

            conguaglio = v['conguaglio'] if 'conguaglio' in v else False
            pod = v['codice_pod'][0] if 'codice_pod' in v else None
            out.append({'row':row, 'conguaglio':conguaglio, 'pod':pod})

    with open(input_file, 'r') as file:
        for r in json.load(file):
            add_rows(r)

    wb = Workbook()
    ws = wb.active

    ws.append([obj.title for obj in table_values])
    for i, obj in enumerate(table_values, 1):
        if obj.column_width != None:
            ws.column_dimensions[get_column_letter(i)].width = obj.column_width

    last_pod = None
    for i, row in enumerate(out, 2):
        cur_pod = row['pod']
        for j, c in enumerate(row['row'], 1):
            cell = ws.cell(row=i, column=j)
            if last_pod is not None and cur_pod is not None and cur_pod != last_pod:
                cell.border = Border(top=Side(border_style='thin', color='000000'))
            if row['conguaglio']:
                cell.fill = PatternFill(patternType='solid', fgColor='ffff00')
            try:
                cell.number_format = c['number_format']
                cell.value = c['value']
                if 'fgColor' in c: cell.fill = PatternFill(patternType='solid', fgColor = c['fgColor'])
            except (KeyError, IndexError, ValueError):
                pass
        last_pod = cur_pod

    for row in out_err:
        ws.append(row)

    ws.freeze_panes = ws['A2']

    wb.save(input_file.with_suffix('.xlsx'))
    print(input_file)

in_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent.parent / 'work/letture'

if __name__ == '__main__':
    if in_path.is_dir():
        for f in in_path.rglob('*.json'):
            export_file(f)
    else:
        export_file(f)
